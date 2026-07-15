"""
B站广告账户管理器 - 桌面应用启动器
数据存储: SQLite 本地数据库 (accounts.db)
"""
import os
import sys
import json
import sqlite3
import threading
import time
import uuid
from datetime import datetime
import webview
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def get_app_dir():
    """获取应用所在目录"""
    return os.path.dirname(os.path.abspath(__file__))


def normalize_date(value):
    """将各种日期格式统一转为 YYYY-MM-DD"""
    import re
    if not value:
        return ""
    s = str(value).strip()
    # 已经是 YYYY-MM-DD
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        return s
    # YYYY/MM/DD
    if re.match(r"^\d{4}/\d{2}/\d{2}$", s):
        return s.replace("/", "-")
    # YYYY.MM.DD
    if re.match(r"^\d{4}\.\d{2}\.\d{2}$", s):
        return s.replace(".", "-")
    # 8位纯数字: 20260701
    if re.match(r"^\d{8}$", s):
        return f"{s[:4]}-{s[4:6]}-{s[6:8]}"
    # 6位纯数字: 260701 → 2026-07-01
    if re.match(r"^\d{6}$", s):
        return f"20{s[:2]}-{s[2:4]}-{s[4:6]}"
    # 已经是其他格式的，原样返回
    return s


def sanitize_past_periods(raw_past, current_start):
    """
    清理往期投放数据：
    - 统一分隔符（换行/中文分号 → ;）
    - 识别 "260709-" 格式（无结束日期=仍在投），提取为开始日期
    - 返回 (清理后的past_periods, 提取到的start_date)
    """
    import re
    if not raw_past:
        return "", current_start

    # 统一分隔符
    s = str(raw_past).strip().replace("；", ";").replace("\r", "").replace("\n", ";")
    parts = [p.strip() for p in s.split(";") if p.strip()]

    ongoing_start = None
    clean_parts = []

    for p in parts:
        # 匹配 "YYMMDD-" 格式（无结束日期=仍在投）
        if re.match(r"^\d{6}-$", p):
            short = p[:6]
            if not ongoing_start:
                ongoing_start = normalize_date(short)
        # 匹配 "YYYYMMDD-" 格式
        elif re.match(r"^\d{8}-$", p):
            if not ongoing_start:
                ongoing_start = normalize_date(p[:8])
        else:
            clean_parts.append(p)

    # 如果有提取到的开始日期，且当前没有开始日期，则使用提取的
    new_start = ongoing_start if ongoing_start and not current_start else current_start
    new_past = "; ".join(clean_parts)

    return new_past, new_start


def get_status_label_from_dates(start_date, pause_date):
    """根据日期计算状态"""
    # 两个日期都为空 → 已暂停
    if not start_date and not pause_date:
        return "已暂停"

    now = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    try:
        start = datetime.strptime(start_date, "%Y-%m-%d") if start_date else None
    except ValueError:
        start = None

    if start and start > now:
        return "待投放"

    if pause_date:
        try:
            pause = datetime.strptime(pause_date, "%Y-%m-%d")
        except ValueError:
            return "投放中"
        if now >= pause:
            return "已暂停"
        return "投放中"

    return "投放中"


class Database:
    """SQLite 数据库管理"""

    def __init__(self, db_path):
        self.db_path = db_path
        self._local = threading.local()
        self._init_db()

    def _get_conn(self):
        """获取当前线程的数据库连接"""
        if not hasattr(self._local, "conn") or self._local.conn is None:
            self._local.conn = sqlite3.connect(self.db_path)
            self._local.conn.row_factory = sqlite3.Row
            self._local.conn.execute("PRAGMA journal_mode=WAL")
            self._local.conn.execute("PRAGMA foreign_keys=ON")
        return self._local.conn

    def _init_db(self):
        """初始化数据库表结构"""
        conn = self._get_conn()
        conn.execute("""
            CREATE TABLE IF NOT EXISTS accounts (
                id TEXT PRIMARY KEY,
                product_name TEXT NOT NULL,
                account_id TEXT NOT NULL,
                account_name TEXT NOT NULL,
                start_date TEXT,
                pause_date TEXT,
                remarks TEXT DEFAULT '',
                past_periods TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT
            )
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_accounts_product
            ON accounts(product_name)
        """)
        conn.commit()

    def list_all(self):
        """返回所有账户，按创建时间倒序"""
        conn = self._get_conn()
        rows = conn.execute(
            "SELECT * FROM accounts ORDER BY created_at DESC"
        ).fetchall()
        return [dict(r) for r in rows]

    def add(self, account):
        """添加账户，account 为 dict"""
        conn = self._get_conn()
        conn.execute("""
            INSERT INTO accounts (id, product_name, account_id, account_name,
                                  start_date, pause_date, remarks, past_periods, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(account_id) DO UPDATE SET
                product_name = excluded.product_name,
                account_name = excluded.account_name,
                start_date = excluded.start_date,
                pause_date = excluded.pause_date,
                remarks = excluded.remarks,
                past_periods = excluded.past_periods,
                updated_at = excluded.updated_at
        """, (
            account["id"],
            account["productName"],
            account["accountId"],
            account["accountName"],
            account["startDate"],
            account.get("pauseDate"),
            account.get("remarks", ""),
            account.get("pastPeriods", ""),
            account["createdAt"],
            account.get("updatedAt"),
        ))
        conn.commit()
        return "ok"

    def update(self, account):
        """更新账户"""
        conn = self._get_conn()
        conn.execute("""
            UPDATE accounts
            SET product_name = ?, account_id = ?, account_name = ?,
                start_date = ?, pause_date = ?, remarks = ?, past_periods = ?, updated_at = ?
            WHERE id = ?
        """, (
            account["productName"],
            account["accountId"],
            account["accountName"],
            account["startDate"],
            account.get("pauseDate"),
            account.get("remarks", ""),
            account.get("pastPeriods", ""),
            account.get("updatedAt"),
            account["id"],
        ))
        conn.commit()
        return "ok"

    def delete(self, account_id):
        """删除账户"""
        conn = self._get_conn()
        conn.execute("DELETE FROM accounts WHERE id = ?", (account_id,))
        conn.commit()
        return "ok"

    def import_batch(self, accounts):
        """批量导入，已存在的 account_id 则合并更新"""
        conn = self._get_conn()
        count = 0
        for a in accounts:
            if not a.get("productName") or not a.get("accountId"):
                continue
            conn.execute("""
                INSERT INTO accounts (id, product_name, account_id,
                    account_name, start_date, pause_date, remarks, past_periods, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(account_id) DO UPDATE SET
                    product_name = COALESCE(NULLIF(excluded.product_name, ''), product_name),
                    account_name = COALESCE(NULLIF(excluded.account_name, ''), account_name),
                    start_date = COALESCE(excluded.start_date, start_date),
                    pause_date = COALESCE(excluded.pause_date, pause_date),
                    remarks = CASE WHEN excluded.remarks != '' THEN excluded.remarks ELSE remarks END,
                    past_periods = excluded.past_periods,
                    updated_at = excluded.updated_at
            """, (
                a["id"],
                a["productName"],
                a["accountId"],
                a["accountName"],
                a["startDate"],
                a.get("pauseDate"),
                a.get("remarks", ""),
                a.get("pastPeriods", ""),
                a.get("createdAt", datetime.now().isoformat()),
                a.get("updatedAt"),
            ))
            count += 1
        conn.commit()
        return count

    def close(self):
        if hasattr(self._local, "conn") and self._local.conn:
            self._local.conn.close()
            self._local.conn = None


class BilibiliAccountAPI:
    """
    暴露给前端调用的 Python API
    所有数据操作走 SQLite
    """

    def __init__(self):
        self.app_dir = get_app_dir()
        self.db_path = os.path.join(self.app_dir, "accounts.db")
        self.backup_path = os.path.join(self.app_dir, "accounts_backup.xlsx")
        self.template_path = os.path.join(self.app_dir, "导入模板.xlsx")
        self.backup_dir = os.path.join(self.app_dir, "backups")
        self.db = Database(self.db_path)

    def get_version(self):
        return "2.0.0"

    def download_template(self) -> str:
        """
        弹出保存对话框，将模板文件保存到用户指定路径
        返回 "ok" 或 "cancelled" 或错误信息
        """
        try:
            import tkinter as tk
            from tkinter import filedialog
            import shutil

            root = tk.Tk()
            root.withdraw()
            root.attributes("-topmost", True)

            filepath = filedialog.asksaveasfilename(
                title="保存导入模板",
                defaultextension=".xlsx",
                filetypes=[("Excel 文件", "*.xlsx")],
                initialfile="B站账户导入模板.xlsx",
            )
            root.destroy()

            if not filepath:
                return "cancelled"

            shutil.copy2(self.template_path, filepath)
            return "ok"
        except Exception as e:
            return f"error: {str(e)}"

    # ---- 数据库 CRUD ----

    def list_accounts(self) -> str:
        """获取所有账户，返回 JSON 字符串"""
        try:
            accounts = self.db.list_all()
            # 将数据库字段映射为前端字段
            result = []
            for a in accounts:
                result.append({
                    "id": a["id"],
                    "productName": a["product_name"],
                    "accountId": a["account_id"],
                    "accountName": a["account_name"],
                    "startDate": a["start_date"],
                    "pauseDate": a["pause_date"],
                    "remarks": a["remarks"] or "",
                    "pastPeriods": a["past_periods"] or "",
                    "createdAt": a["created_at"],
                    "updatedAt": a["updated_at"],
                })
            return json.dumps(result, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"error": str(e)})

    def add_account(self, account_json: str) -> str:
        """添加账户，返回新账户的 JSON"""
        try:
            data = json.loads(account_json)
            now = datetime.now().isoformat()
            raw_past = data.get("pastPeriods", "").strip().replace("；", ";").replace("\r", "").replace("\n", ";")
            clean_past, extracted_start = sanitize_past_periods(raw_past, data.get("startDate"))
            account = {
                "id": "acc_" + uuid.uuid4().hex[:12],
                "productName": data.get("productName", "").strip(),
                "accountId": data.get("accountId", "").strip(),
                "accountName": data.get("accountName", "").strip(),
                "startDate": normalize_date(extracted_start) if extracted_start else (data.get("startDate") or None),
                "pauseDate": data.get("pauseDate") or None,
                "remarks": data.get("remarks", "").strip(),
                "pastPeriods": clean_past,
                "createdAt": now,
                "updatedAt": None,
            }
            self.db.add(account)
            self._sync_excel()
            return json.dumps(account, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"error": str(e)})

    def update_account(self, account_json: str) -> str:
        """更新账户"""
        try:
            data = json.loads(account_json)
            account = {
                "id": data["id"],
                "productName": data.get("productName", "").strip(),
                "accountId": data.get("accountId", "").strip(),
                "accountName": data.get("accountName", "").strip(),
                "startDate": data.get("startDate") or None,
                "pauseDate": data.get("pauseDate") or None,
                "remarks": data.get("remarks", "").strip(),
                "pastPeriods": data.get("pastPeriods", "").strip().replace("；", ";").replace("\r", "").replace("\n", ";"),
                "updatedAt": datetime.now().isoformat(),
            }
            self.db.update(account)
            self._sync_excel()
            return "ok"
        except Exception as e:
            return f"error: {str(e)}"

    def delete_account(self, account_id: str) -> str:
        """删除账户"""
        try:
            self.db.delete(account_id)
            self._sync_excel()
            return "ok"
        except Exception as e:
            return f"error: {str(e)}"

    def import_accounts(self, accounts_json: str) -> str:
        """批量导入"""
        try:
            data = json.loads(accounts_json)
            accounts = []
            now = datetime.now().isoformat()
            for item in data:
                if not item.get("productName") or not item.get("accountId"):
                    continue
                raw_past = item.get("pastPeriods", item.get("往期投放", ""))
                clean_past, extracted_start = sanitize_past_periods(raw_past, item.get("startDate"))
                accounts.append({
                    "id": item.get("id", "acc_" + uuid.uuid4().hex[:12]),
                    "productName": item.get("productName", "").strip(),
                    "accountId": item.get("accountId", "").strip(),
                    "accountName": item.get("accountName", "").strip(),
                    "startDate": normalize_date(extracted_start) if extracted_start else normalize_date(item.get("startDate", "")),
                    "pauseDate": normalize_date(item.get("pauseDate")) if item.get("pauseDate") else None,
                    "remarks": item.get("remarks", item.get("备注", "")).strip(),
                    "pastPeriods": clean_past,
                    "createdAt": item.get("createdAt", now),
                    "updatedAt": item.get("updatedAt"),
                })
            count = self.db.import_batch(accounts)
            self._sync_excel()
            return str(count)
        except Exception as e:
            return f"error: {str(e)}"

    def rename_product(self, data_json: str) -> str:
        """批量修改产品名"""
        try:
            data = json.loads(data_json)
            old_name = data.get("oldName", "").strip()
            new_name = data.get("newName", "").strip()
            if not old_name or not new_name:
                return "error: 名称不能为空"
            conn = self.db._get_conn()
            affected = conn.execute(
                "UPDATE accounts SET product_name = ?, updated_at = ? WHERE product_name = ?",
                (new_name, datetime.now().isoformat(), old_name)
            ).rowcount
            conn.commit()
            self._sync_excel()
            return "ok"
        except Exception as e:
            return f"error: {str(e)}"

    def archive_period(self, account_id: str) -> str:
        """将当前投放时段归档到往期投放，清空开始/暂停日期"""
        try:
            conn = self.db._get_conn()
            row = conn.execute(
                "SELECT start_date, pause_date, past_periods FROM accounts WHERE id = ?",
                (account_id,)
            ).fetchone()
            if not row or not row["start_date"] or not row["pause_date"]:
                return "error: 无时段可归档"

            # 格式化: 2026-07-01 → 260701
            def to_short(d):
                return d.replace("-", "")[2:] if d else ""

            period = f"{to_short(row['start_date'])}-{to_short(row['pause_date'])}"
            existing = (row["past_periods"] or "").strip()
            new_past = (existing + "; " + period).strip("; ")

            conn.execute(
                "UPDATE accounts SET past_periods = ?, start_date = NULL, pause_date = NULL, updated_at = ? WHERE id = ?",
                (new_past, datetime.now().isoformat(), account_id)
            )
            conn.commit()
            self._sync_excel()
            return "ok"
        except Exception as e:
            return f"error: {str(e)}"

    # ---- Excel 备份（从 SQLite 读取） ----

    def _sync_excel(self):
        """从 SQLite 读取数据写入 Excel"""
        try:
            accounts = self.db.list_all()
            self._write_excel(accounts)
        except Exception as e:
            print(f"Excel 同步失败: {e}")

    def save_to_excel(self, _=None) -> str:
        """手动触发 Excel 同步（兼容旧接口）"""
        try:
            accounts = self.db.list_all()
            self._write_excel(accounts)
            return "ok"
        except Exception as e:
            return f"error: {str(e)}"

    def create_snapshot_backup(self) -> str:
        """创建带时间戳的快照备份，最多保留10份"""
        try:
            os.makedirs(self.backup_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"accounts_backup_{timestamp}.xlsx"
            filepath = os.path.join(self.backup_dir, filename)

            accounts = self.db.list_all()
            self._write_excel(accounts)
            import shutil
            shutil.copy2(self.backup_path, filepath)

            backups = sorted(
                [f for f in os.listdir(self.backup_dir) if f.endswith('.xlsx')],
                reverse=True
            )
            for old in backups[10:]:
                os.remove(os.path.join(self.backup_dir, old))

            return json.dumps({"ok": True, "file": filename, "count": len(backups[:10])})
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

    def restore_from_backup(self) -> str:
        """弹出打开文件对话框，选择备份文件恢复数据"""
        try:
            import tkinter as tk
            from tkinter import filedialog

            root = tk.Tk()
            root.withdraw()
            root.attributes("-topmost", True)

            filepath = filedialog.askopenfilename(
                title="选择备份文件恢复",
                filetypes=[("Excel 文件", "*.xlsx")],
                initialdir=self.backup_dir,
            )
            root.destroy()

            if not filepath:
                return "cancelled"

            from openpyxl import load_workbook
            wb = load_workbook(filepath, read_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(min_row=1, max_row=1)
            header_row = [cell.value for cell in next(rows_iter)]
            col_map = {str(h).strip(): i for i, h in enumerate(header_row) if h}

            accounts = []
            now = datetime.now().isoformat()
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[col_map.get("产品名", 0) if col_map else 0]:
                    continue
                raw_past = str(row[col_map.get("往期投放", 8)]).strip() if col_map.get("往期投放") and len(row) > 8 and row[8] else ""
                clean_past, extracted_start = sanitize_past_periods(raw_past, row[col_map.get("投放开始", 4)] if col_map.get("投放开始") and row[col_map.get("投放开始")] else "")
                accounts.append({
                    "id": "acc_" + uuid.uuid4().hex[:12],
                    "productName": str(row[col_map.get("产品名", 0)]).strip() if col_map else "",
                    "accountId": str(row[col_map.get("账户ID", 1)]).strip() if col_map else "",
                    "accountName": str(row[col_map.get("账户名", 2)]).strip() if col_map else "",
                    "startDate": normalize_date(extracted_start) if extracted_start else normalize_date(str(row[col_map.get("投放开始", 4)]) if col_map.get("投放开始") and row[col_map["投放开始"]] else ""),
                    "pauseDate": normalize_date(str(row[col_map.get("投放暂停", 5)])) if col_map.get("投放暂停") and row[col_map["投放暂停"]] else None,
                    "pastPeriods": clean_past,
                    "remarks": str(row[col_map.get("备注", 7)]).strip() if col_map.get("备注") and len(row) > 7 and row[7] else "",
                    "createdAt": now,
                    "updatedAt": None,
                })

            wb.close()
            self.db.import_batch(accounts)
            self._sync_excel()
            return json.dumps({"ok": True, "count": len(accounts)})
        except Exception as e:
            return json.dumps({"ok": False, "error": str(e)})

    def export_to_xlsx(self) -> str:
        """弹出保存对话框，导出为 Excel"""
        try:
            import tkinter as tk
            from tkinter import filedialog
            import shutil

            # 先生成最新备份
            accounts = self.db.list_all()
            self._write_excel(accounts)

            root = tk.Tk()
            root.withdraw()
            root.attributes("-topmost", True)

            filepath = filedialog.asksaveasfilename(
                title="导出账户数据",
                defaultextension=".xlsx",
                filetypes=[("Excel 文件", "*.xlsx")],
                initialfile=f"accounts_export_{datetime.now().strftime('%Y%m%d')}.xlsx",
            )
            root.destroy()

            if not filepath:
                return "cancelled"

            shutil.copy2(self.backup_path, filepath)
            return "ok"
        except Exception as e:
            return f"error: {str(e)}"

    def _write_excel(self, accounts):
        """将账户列表写入 Excel 文件"""
        wb = Workbook()
        ws = wb.active
        ws.title = "B站广告账户"

        header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="FB7299", end_color="FB7299", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        body_font = Font(name="微软雅黑", size=10)
        body_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="D0D0D8"),
            right=Side(style="thin", color="D0D0D8"),
            top=Side(style="thin", color="D0D0D8"),
            bottom=Side(style="thin", color="D0D0D8"),
        )

        status_fills = {
            "投放中": PatternFill(start_color="E8F8E8", end_color="E8F8E8", fill_type="solid"),
            "已暂停": PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
            "待投放": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
            "已暂停": PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
        }
        status_fonts = {
            "投放中": Font(name="微软雅黑", size=10, color="27AE60"),
            "已暂停": Font(name="微软雅黑", size=10, color="E67E22"),
            "待投放": Font(name="微软雅黑", size=10, color="1976D2"),
            "已暂停": Font(name="微软雅黑", size=10, color="999999"),
        }

        headers = ["序号", "产品名", "账户ID", "账户名", "投放开始", "投放暂停", "状态", "备注", "往期投放", "创建时间", "更新时间"]
        col_widths = [8, 22, 18, 22, 16, 16, 12, 24, 26, 20, 20]

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.freeze_panes = "A2"

        for row_idx, a in enumerate(accounts, 1):
            data_idx = row_idx + 1
            status = get_status_label_from_dates(a.get("start_date", ""), a.get("pause_date"))

            row_data = [
                row_idx,
                a.get("product_name", ""),
                a.get("account_id", ""),
                a.get("account_name", ""),
                a.get("start_date", ""),
                a.get("pause_date") or "",
                status,
                a.get("remarks") or "",
                a.get("past_periods") or "",
                a.get("created_at", ""),
                a.get("updated_at") or "",
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=data_idx, column=col_idx, value=value)
                cell.font = body_font
                cell.alignment = body_alignment
                cell.border = thin_border

                if col_idx == 7:
                    fill = status_fills.get(value)
                    font = status_fonts.get(value)
                    if fill:
                        cell.fill = fill
                    if font:
                        cell.font = font

            ws.row_dimensions[data_idx].height = 22

        ws.row_dimensions[1].height = 28
        wb.save(self.backup_path)

    # ---- Excel 数据迁移（旧数据导入 SQLite） ----

    def load_from_excel(self) -> str:
        """从 Excel 读取数据并导入 SQLite（用于数据迁移，不覆盖已有数据）"""
        try:
            from openpyxl import load_workbook

            if not os.path.exists(self.backup_path):
                return json.dumps({"imported": 0, "accounts": []})

            wb = load_workbook(self.backup_path, read_only=True)
            ws = wb.active
            accounts = []
            now = datetime.now().isoformat()

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[1]:
                    continue

                # 兼容旧版 Excel
                col_count = len(row)
                if col_count >= 11:
                    # 最新格式: 序号/产品名/账户ID/账户名/开始/暂停/状态/备注/往期投放/创建时间/更新时间
                    accounts.append({
                        "id": "acc_" + uuid.uuid4().hex[:12],
                        "productName": str(row[1]) if row[1] else "",
                        "accountId": str(row[2]) if row[2] else "",
                        "accountName": str(row[3]) if row[3] else "",
                        "startDate": str(row[4]) if row[4] else "",
                        "pauseDate": str(row[5]) if row[5] and str(row[5]).strip() else None,
                        "remarks": str(row[7]) if row[7] else "",
                        "pastPeriods": str(row[8]) if row[8] else "",
                        "createdAt": str(row[9]) if row[9] else now,
                        "updatedAt": str(row[10]) if row[10] else None,
                    })
                elif col_count >= 10:
                    # 旧格式(有备注): 序号/产品名/账户ID/账户名/开始/暂停/状态/备注/创建时间/更新时间
                    accounts.append({
                        "id": "acc_" + uuid.uuid4().hex[:12],
                        "productName": str(row[1]) if row[1] else "",
                        "accountId": str(row[2]) if row[2] else "",
                        "accountName": str(row[3]) if row[3] else "",
                        "startDate": str(row[4]) if row[4] else "",
                        "pauseDate": str(row[5]) if row[5] and str(row[5]).strip() else None,
                        "remarks": str(row[7]) if row[7] else "",
                        "pastPeriods": "",
                        "createdAt": str(row[8]) if row[8] else now,
                        "updatedAt": str(row[9]) if row[9] else None,
                    })
                else:
                    # 旧格式（无备注列）
                    accounts.append({
                        "id": "acc_" + uuid.uuid4().hex[:12],
                        "productName": str(row[1]) if row[1] else "",
                        "accountId": str(row[2]) if row[2] else "",
                        "accountName": str(row[3]) if row[3] else "",
                        "startDate": str(row[4]) if row[4] else "",
                        "pauseDate": str(row[5]) if row[5] and str(row[5]).strip() else None,
                        "remarks": "",
                        "createdAt": str(row[7]) if row[7] else now,
                        "updatedAt": str(row[8]) if row[8] else None,
                    })

            wb.close()

            if accounts:
                self.db.import_batch(accounts)

            return json.dumps({
                "imported": len(accounts),
                "accounts": self.db.list_all()
            }, ensure_ascii=False, default=str)

        except Exception as e:
            print(f"从 Excel 迁移数据失败: {e}")
            return json.dumps({"imported": 0, "accounts": []})


def _write_template(filepath):
    """生成空白导入模板 Excel 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "导入模板"

    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FB7299", end_color="FB7299", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    body_font = Font(name="微软雅黑", size=10)
    body_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D0D0D8"),
        right=Side(style="thin", color="D0D0D8"),
        top=Side(style="thin", color="D0D0D8"),
        bottom=Side(style="thin", color="D0D0D8"),
    )

    headers = ["产品名", "账户ID", "账户名", "往期投放", "投放开始", "投放暂停", "备注"]
    col_widths = [22, 18, 24, 26, 16, 16, 24]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    example = ["原神", "1234567", "原神主账户", "260101-260202; 260301-260415", "260701", "260801", "示例备注"]
    for col_idx, value in enumerate(example, 1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.font = body_font
        cell.alignment = body_alignment
        cell.border = thin_border

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A2"
    wb.save(filepath)


def main():
    app_dir = get_app_dir()
    html_path = os.path.join(app_dir, "index.html")

    if not os.path.exists(html_path):
        print(f"错误：找不到 {html_path}")
        sys.exit(1)

    api = BilibiliAccountAPI()

    # 启动时生成导入模板文件
    template_path = os.path.join(app_dir, "导入模板.xlsx")
    _write_template(template_path)

    # HTTP 服务器（静态文件 + API 端点）
    import http.server
    import socketserver

    PORT = 18081
    os.chdir(app_dir)

    class APIHandler(http.server.SimpleHTTPRequestHandler):
        """自定义请求处理器，静态文件 + REST API"""

        def _send_json(self, data, status=200):
            self.send_response(status)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Access-Control-Allow-Origin", "*")
            self.send_header("Cache-Control", "no-cache, no-store, must-revalidate")
            self.end_headers()
            self.wfile.write(data.encode("utf-8"))

        def end_headers(self):
            self.send_header("Cache-Control", "no-cache, no-store, must-revalidate")
            super().end_headers()

        def _read_body(self):
            length = int(self.headers.get("Content-Length", 0))
            return self.rfile.read(length).decode("utf-8") if length > 0 else ""

        def do_GET(self):
            if self.path == "/api/accounts":
                self._send_json(api.list_accounts())
            elif self.path == "/api/template":
                self._send_template()
            else:
                super().do_GET()

        def do_POST(self):
            if self.path == "/api/accounts":
                body = self._read_body()
                result = api.add_account(body)
                self._send_json(result, 201)
            elif self.path == "/api/accounts/import":
                body = self._read_body()
                result = api.import_accounts(body)
                self._send_json(json.dumps({"count": int(result) if result.isdigit() else 0}))
            elif self.path == "/api/import-xlsx":
                self._handle_xlsx_import()
            elif self.path == "/api/rename-product":
                body = self._read_body()
                result = api.rename_product(body)
                self._send_json(json.dumps({"ok": result == "ok"}))
            elif self.path.startswith("/api/archive/"):
                account_id = self.path.split("/")[-1]
                result = api.archive_period(account_id)
                self._send_json(json.dumps({"ok": result == "ok"}))
            elif self.path == "/api/backup":
                result = api.create_snapshot_backup()
                self._send_json(result)
            elif self.path == "/api/restore":
                result = api.restore_from_backup()
                self._send_json(result)
            else:
                self.send_response(404)
                self.end_headers()

        def do_PUT(self):
            if self.path == "/api/accounts":
                body = self._read_body()
                result = api.update_account(body)
                if result == "ok":
                    self._send_json('"ok"')
                else:
                    self._send_json(json.dumps({"error": result}), 500)
            else:
                self.send_response(404)
                self.end_headers()

        def do_DELETE(self):
            if self.path.startswith("/api/accounts/"):
                account_id = self.path.split("/")[-1]
                result = api.delete_account(account_id)
                if result == "ok":
                    self._send_json('"ok"')
                else:
                    self._send_json(json.dumps({"error": result}), 500)
            else:
                self.send_response(404)
                self.end_headers()

        def do_OPTIONS(self):
            self.send_response(204)
            self.send_header("Access-Control-Allow-Origin", "*")
            self.send_header("Access-Control-Allow-Methods", "GET, POST, PUT, DELETE, OPTIONS")
            self.send_header("Access-Control-Allow-Headers", "Content-Type")
            self.end_headers()

        def _send_template(self):
            """生成空白导入模板 Excel"""
            from io import BytesIO
            wb = Workbook()
            ws = wb.active
            ws.title = "导入模板"

            header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="FB7299", end_color="FB7299", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            body_font = Font(name="微软雅黑", size=10)
            body_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin", color="D0D0D8"),
                right=Side(style="thin", color="D0D0D8"),
                top=Side(style="thin", color="D0D0D8"),
                bottom=Side(style="thin", color="D0D0D8"),
            )

            headers = ["产品名", "账户ID", "账户名", "投放开始", "投放暂停", "备注"]
            col_widths = [22, 18, 24, 16, 16, 24]

            for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            # 添加一行示例数据
            example = ["原神", "1234567", "原神主账户", "260101-260202; 260301-260415", "260701", "260801", "示例备注"]
            for col_idx, value in enumerate(example, 1):
                cell = ws.cell(row=2, column=col_idx, value=value)
                cell.font = body_font
                cell.alignment = body_alignment
                cell.border = thin_border

            ws.row_dimensions[1].height = 28
            ws.row_dimensions[2].height = 22
            ws.freeze_panes = "A2"

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", 'attachment; filename="B站账户导入模板.xlsx"')
            self.end_headers()
            self.wfile.write(output.read())

        def _handle_xlsx_import(self):
            """解析上传的 Excel 文件并导入"""
            from io import BytesIO
            from openpyxl import load_workbook

            content_type = self.headers.get("Content-Type", "")
            if "multipart/form-data" not in content_type:
                self._send_json(json.dumps({"error": "不支持的上传格式"}), 400)
                return

            # 简单 multipart 解析
            body = self.rfile.read(int(self.headers.get("Content-Length", 0)))
            boundary = content_type.split("boundary=")[-1].encode()
            parts = body.split(b"--" + boundary)

            for part in parts:
                if b"filename=" in part:
                    # 找到文件内容（header 后面的空行之后）
                    idx = part.find(b"\r\n\r\n")
                    if idx >= 0:
                        file_data = part[idx + 4:]
                        # 去掉结尾的 \r\n
                        if file_data.endswith(b"\r\n"):
                            file_data = file_data[:-2]

                        try:
                            wb = load_workbook(BytesIO(file_data), read_only=True)
                            ws = wb.active
                            accounts = []
                            now = datetime.now().isoformat()

                            header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                            col_map = {}
                            for i, h in enumerate(header_row):
                                if h:
                                    col_map[str(h).strip()] = i

                            for row in ws.iter_rows(min_row=2, values_only=True):
                                if not row or not row[col_map.get("产品名", 1) if col_map else 1]:
                                    continue
                                raw_past = str(row[col_map.get("往期投放", 3)]).strip() if col_map and len(header_row) > 3 and row[col_map.get("往期投放", 3)] else ""
                                clean_past, extracted_start = sanitize_past_periods(raw_past, row[col_map.get("投放开始", 4)] if col_map and row[col_map.get("投放开始", 4)] else "")
                                accounts.append({
                                    "id": "acc_" + uuid.uuid4().hex[:12],
                                    "productName": str(row[col_map.get("产品名", 0)]).strip() if col_map else "",
                                    "accountId": str(row[col_map.get("账户ID", 1)]).strip() if col_map else "",
                                    "accountName": str(row[col_map.get("账户名", 2)]).strip() if col_map else "",
                                    "startDate": normalize_date(extracted_start) if extracted_start else normalize_date(row[col_map.get("投放开始", 4)] if col_map else ""),
                                    "pauseDate": normalize_date(row[col_map.get("投放暂停", 5)]) if col_map and row[col_map.get("投放暂停", 5)] else None,
                                    "pastPeriods": clean_past,
                                    "remarks": str(row[col_map.get("备注", 6)]).strip() if col_map and len(header_row) > 6 and row[col_map.get("备注", 6)] else "",
                                    "createdAt": now,
                                    "updatedAt": None,
                                })

                            wb.close()
                            count = api.db.import_batch(accounts)
                            api._sync_excel()
                            self._send_json(json.dumps({"count": count}))
                            return
                        except Exception as ex:
                            self._send_json(json.dumps({"error": str(ex)}), 500)
                            return

            self._send_json(json.dumps({"error": "未找到文件内容"}), 400)

        def log_message(self, format, *args):
            if "404" in str(args) and "favicon" in str(args):
                return
            super().log_message(format, *args)

    with socketserver.TCPServer(("127.0.0.1", PORT), APIHandler) as httpd:
        server_thread = threading.Thread(target=httpd.serve_forever, daemon=True)
        server_thread.start()

        # 每小时自动备份
        def auto_backup():
            while True:
                time.sleep(3600)
                try:
                    api.create_snapshot_backup()
                    print(f"[自动备份] {datetime.now().strftime('%H:%M:%S')}")
                except Exception as e:
                    print(f"[自动备份失败] {e}")
        backup_thread = threading.Thread(target=auto_backup, daemon=True)
        backup_thread.start()

        window = webview.create_window(
            title="B站广告账户管理器",
            url=f"http://127.0.0.1:{PORT}/index.html",
            js_api=api,
            width=1200,
            height=800,
            min_size=(900, 600),
            confirm_close=True,
            text_select=True,
        )

        webview.start(debug=False, http_server=False)

        httpd.shutdown()
        api.db.close()


if __name__ == "__main__":
    main()
